
from openpyxl import load_workbook
from matplotlib import pyplot

def getvalue(x): return x.value

wb = load_workbook("data_analysis_lab.xlsx")

ws = wb["Data"]

columnA = ws["A"][1:]
columnC = ws["C"][1:]
columnD = ws["D"][1:]

pyplot.plot(list(map(getvalue, columnA)), list(map(getvalue, columnC)))
pyplot.plot(list(map(getvalue, columnA)), list(map(getvalue, columnD)))

pyplot.show()