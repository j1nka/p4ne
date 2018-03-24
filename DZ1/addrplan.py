import glob
import re
import ipaddress
from openpyxl import load_workbook
from openpyxl.styles import Alignment


#определяем форматирование excel
alignment=Alignment(horizontal='center',
vertical='center',
text_rotation=0,
wrap_text=True,
shrink_to_fit=False,
indent=0)

#подготавливаем Excel для внесения данных
wb = load_workbook("addrplan.xlsx")
ws = wb["AddressPlan"]
ws["A1"] = "№"
ws.column_dimensions["A"].width = 6
ws["A1"].alignment = alignment
ws["B1"] = "IP address"
ws.column_dimensions["B"].width = 20
ws["B1"].alignment = alignment
ws["C1"] = "Mask"
ws.column_dimensions["C"].width = 20
ws["C1"].alignment = alignment
counter = 2 #пригодится для записи в excel-file

#создаем функцию, которая будет вытаскивать ip и маски
#при помощи регулярных выражений
def getipfromline(line):
    pat = "(ip address) ((?:[0-9]{1,3}[\.]){1,3}[0-9]{1,3}) ((?:[0-9]{1,3}[\.]){3}[0-9]{1,3})"
    m = re.match(pat, line.strip().lower())
    if m:
        ip = ipaddress.IPv4Interface(m.group(2) + "/" + m.group(3))
        return ip
    else:
        return -1
def sortfunc(data):
    return [int(data.netmask), int(data.ip)]

result = {}

#читаем все конфигурационные файлы и добавляем ip в словарь

for fname in glob.glob("config_files\*.txt"):
    for line in open(fname):
       iptry = getipfromline(line)
       if iptry != -1:
            if fname in result.keys():
                result[fname].append(iptry)
            else:
                result[fname] = [iptry]

#чистим получившиеся списки от повторов

for key in result.keys():
    result[key] = list(set(result[key]))

#выводим на экран в упорядоченном виде
#сперва "по маске", потом "по ip"
#сортировка осуществляется отдельно для каждого файла
#добавляем данные в excel-файл
nums = 0

for key in result.keys():
    print(key)
    ws["A"+str(counter)] = key
    ws["A"+str(counter)].alignment = alignment
    ws.merge_cells(start_row=counter, start_column=1, end_row=counter, end_column=3)
    counter += 1
    nums+=1
    for data in sorted(result[key], key=sortfunc):
        #print(data)
        print("IP address: " + str(data.ip) + "; Mask: " + str(data.netmask))
        ws["A"+str(counter)] = counter-1-nums
        ws["A" + str(counter)].alignment = alignment
        ws["B"+str(counter)] = str(data.ip)
        ws["B"+str(counter)].alignment = alignment
        ws["C"+str(counter)] = str(data.netmask)
        ws["C"+str(counter)].alignment = alignment
        counter += 1

wb.save("addrplan.xlsx")